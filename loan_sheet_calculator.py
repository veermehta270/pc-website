import os
import csv
import pandas as pd
from tkinter import Tk, filedialog, Label, Entry, Button, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from tkinter import filedialog
from tkinter import *
from PIL import Image, ImageTk
import winsound
from tkinter.ttk import Progressbar
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from pprint import pprint
from tkinter import font
import tkinter as tk
import zipfile



def read_csv(filename):
    if filename.lower().endswith('.csv'):
        with open(filename, newline='', encoding='utf-8') as csvfile:
            return list(csv.reader(csvfile))
    elif filename.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(filename)
        return [df.columns.tolist()] + df.values.tolist()
    else:
        raise ValueError("Unsupported file type. Please upload a .csv or .xlsx/.xls file.")





BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ref_path = os.path.join(BASE_DIR, 'AOR.csv')

aor=read_csv(ref_path)


#border
thick_border = Border(
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# Fill color (light yellow)
fill_yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

# Thin border around cell
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Bold text font
bold_font = Font(bold=True)

# Center alignment
center_align = Alignment(horizontal='center', vertical='center')



def extract(filename):
    add = ""
    for elem in filename:
        if elem == '_':
            break
        else:
            add += elem

    sheet_map = {
        "Transactions": f"{add}_transactions.csv",
        "Large Transactions": f"{add}_large.csv",
    }

    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)

    for sheet_name, output_csv in sheet_map.items():
        df = pd.read_excel(filename, sheet_name=sheet_name)
        for col in df.select_dtypes(include=['float', 'int']).columns:
            if (df[col] % 1 == 0).all():
                df[col] = df[col].astype(int)
        output_csv = os.path.basename(output_csv)
        df.to_csv(os.path.join(output_dir,output_csv),index=False)




def fun(transactions):
    tns = read_csv(transactions)[1:]
    aor_lst = []
    tns_lst = [[],[]]
    res = [[],[]]
    

    for row in aor:
        for elem in row:
            aor_lst.append(round(float(elem)))
    for row in tns:
        if row[4] != '':
            tns_lst[0].append(round(float(row[4])))
            tns_lst[1].append(round(float(row[4])) * 2)

    for elem in aor_lst:
        if (elem in tns_lst[0]) or (elem + 1 in tns_lst[0]) or (elem - 1 in tns_lst[0]):
            res[0].append(elem)
    for elem in aor_lst:
        if (elem in tns_lst[1]) or (elem + 1 in tns_lst[1]) or (elem - 1 in tns_lst[1]):
            res[1].append(elem)

    return res

def freq(num,transactions,date):
    monthly=[]
    biweekly=[]
    rows=read_csv(transactions)
    res_m=0
    res_b=0
    rows=rows[1:]
    for row in rows:
        if row[4]=='':
            continue
        if (round(num)==round(float(row[4])) or round(num)+1==round(float(row[4])) or round(num)-1==round(float(row[4]))) and (round(num) not in biweekly) and (row[1]>=date):
            res_m+=1
            monthly.append(round(float(row[4])))
        elif (round(num)//2==round(float(row[4])) or round(num)//2+1==round(float(row[4])) or round(num)//2-1==round(float(row[4])) and (round(num) not in monthly)) and (row[1]>=date):
            res_b+=1
            biweekly.append(round(float(row[4])))
    return [res_m,res_b]


def fun4(amt, transactions, months, date):
    res = []
    data = fun(transactions)
    for tenor in range(1, 13):
        rate = 0
        while rate <= 0.0525:
            total_int = amt * tenor * rate
            avg_monthly_intrst = total_int / tenor
            avg_principle_repay = amt / tenor
            mo_repay = avg_monthly_intrst + avg_principle_repay
            if ((round(mo_repay) in data[0]) or
                (round(mo_repay) + 1 in data[0]) or
                (round(mo_repay) - 1 in data[0])):
                if tenor in months:
                    res.append([tenor, round(rate, 5), round(mo_repay, 3),'Monthly',freq(round(mo_repay,2),transactions,date)[0]])
            if ((round(mo_repay) in data[1]) or
                (round(mo_repay) + 1 in data[1]) or
                (round(mo_repay) - 1 in data[1])):
                if tenor in months:
                    res.append([tenor, round(rate, 5), round(mo_repay, 3),'Biweekly',freq(round(mo_repay,2),transactions,date)[1]])
            rate += 0.0025
    return res


def get_large_principles(large_transaction):
    rows = read_csv(large_transaction)[1:]
    lst = []
    for row in rows:
        for i in range(5, 11):
            if row[i] != '':
                val = int(float(row[i]))
                if val >= 500000 and val % 500000 == 0:
                    lst.append([row[0], val])
    return lst



def add_days(date_str, days):
    # Parse string into a datetime object
    date_obj = datetime.strptime(date_str, "%Y/%m/%d")
    
    # Add days
    new_date = date_obj + timedelta(days=days)
    
    # Return formatted string
    return new_date.strftime("%Y/%m/%d")




def all_combs(transactions, large_transaction, months):
    end_date=read_csv(transactions)[-1][1]
    res = []
    res2=[]
    for elem in get_large_principles(large_transaction):
        res.append([elem[0], elem[1], fun4(elem[1], transactions, months,elem[0])])

    for elem in res:
        st=elem[0]
        for elem in elem[2]:
            if(elem[3]=='Biweekly'):
                if(add_days(st,14)>end_date):
                    elem.append('No Data')
                else:
                    elem.append('')
            elif(elem[3]=='Monthly'):
                if(add_days(st,30)>end_date):
                    elem.append('No Data')
                else:
                    elem.append('')

    return res


def get_loan_sheet(complete, months):
    extract(complete)
    base_name = os.path.splitext(os.path.basename(complete))[0].split('_')[0]
    transactions = os.path.join('output', f"{base_name}_transactions.csv")
    large = os.path.join('output', f"{base_name}_large.csv")
    data = all_combs(transactions, large, months)

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Data"
    headers = ["Date", "Amount", "Term (Months)", "Interest Rate", "Installment","Monthly/Biweekly","Frequency"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_yellow
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    


    row_index = 2
    for entry in data:
        date, amount, terms = entry
        num_terms = len(terms)
        for i, term in enumerate(terms):
            term_months, rate, installment,monthly_biweekly,frequency,select = term
            ws.cell(row=row_index + i, column=3, value=term_months)
            ws.cell(row=row_index + i, column=4, value=rate)
            ws.cell(row=row_index + i, column=5, value=installment)
            ws.cell(row=row_index + i, column=6, value=monthly_biweekly)
            ws.cell(row=row_index + i, column=7, value=frequency)
            ws.cell(row=row_index + i, column=8, value=select)
        if num_terms > 1:
            ws.merge_cells(start_row=row_index, end_row=row_index + num_terms - 1, start_column=1, end_column=1)
            ws.merge_cells(start_row=row_index, end_row=row_index + num_terms - 1, start_column=2, end_column=2)
        ws.cell(row=row_index, column=1, value=date)
        ws.cell(row=row_index, column=2, value=amount)
        
        
        for col in [1, 2]:
            cell = ws.cell(row=row_index, column=col)
            cell.alignment = Alignment(vertical='center', horizontal='center')
        row_index += num_terms

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    i = 1
    while os.path.exists(f"{base_name}_loan_data_merged_{i}.xlsx"):
        i += 1

    filename = f"{base_name}_loan_data_merged_{i}.xlsx"

    checkbox_col = 8
    ws.cell(row=1, column=checkbox_col,value="Select")
    cell = ws.cell(row=1,column=checkbox_col)
    cell.fill = fill_yellow
    cell.font=bold_font
    cell.border=thin_border
    cell.alignment = center_align

    dv = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
    ws.add_data_validation(dv)

    for r in range(2, row_index):
        cell = ws.cell(row=r, column=checkbox_col)
        if cell.value == "":
            dv.add(cell)
    ws.column_dimensions[get_column_letter(checkbox_col)].width = 12

    
    output_dir='output'
    os.makedirs(output_dir,exist_ok=True)
    i=1
    while os.path.exists(os.path.join(output_dir, f"{base_name}_loan_data_merged_{i}.xlsx")):
        i += 1
    excel_filename = f"{base_name}_loan_data_merged_{i}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)
    wb.save(excel_path)

    csv1_path =  transactions
    csv2_path =  large
    
    
    zip_filename = f"{base_name}_loan_files.zip"
    zip_path = os.path.join(output_dir, zip_filename)
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(excel_path, arcname=excel_filename)
        zipf.write(csv1_path, arcname=transactions)
        zipf.write(csv2_path, arcname=large)

    return zip_filename








print('veer')


    
