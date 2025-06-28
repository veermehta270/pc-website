import os
import pandas as pd
from tkinter import Tk, filedialog, Label, Entry, Button, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from tkinter import filedialog
from tkinter import *
from PIL import Image, ImageTk
import winsound
from tkinter.ttk import Progressbar
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from pprint import pprint
import csv
import tempfile
from openpyxl.styles import PatternFill
from tkinter import ttk
from tkinter import Toplevel, ttk
from tkinter import font




#border
thick_border = Border(
    top=Side(style='thick'),
    bottom=Side(style='thick')
)


# Yellow Color
fill_yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')


# Define fill colors
green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # light green
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')    # very light red






holidays={'2021/12/08', '2022/02/25', '2025/04/01', '2026/05/01', '2022/11/30', '2026/12/30', '2024/11/02', '2028/04/15', '2024/12/08', '2025/12/30', '2023/08/28', '2025/01/29', '2027/05/01', '2024/11/01', '2026/05/27', '2020/01/01', '2022/04/17', '2025/09/05', '2024/05/01', '2023/12/26', '2028/11/30', '2025/04/18', '2026/06/17', '2024/08/26', '2021/06/12', '2028/11/02', '2021/04/04', '2020/11/01', '2028/08/28', '2028/05/25', '2024/12/31', '2025/12/31', '2020/11/30', '2022/01/01', '2023/05/01', '2022/02/28', '2024/03/28', '2024/03/12', '2025/12/24', '2027/12/26', '2023/12/08', '2024/12/25', '2021/08/30', '2025/12/08', '2024/06/18', '2026/08/31', '2021/12/30', '2021/12/25', '2020/12/25', '2022/04/09', '2020/12/31', '2024/06/17', '2020/08/01', '2023/12/30', '2020/04/09', '2025/06/26', '2026/01/16', '2025/08/25', '2025/05/12', '2025/06/07', '2025/11/02', '2025/06/06', '2026/08/21', '2026/12/08', '2027/03/27', '2024/02/25', '2021/04/03', '2024/08/23', '2025/05/01', '2026/01/01', '2020/01/25', '2027/02/25', '2022/12/30', '2022/08/29', '2026/04/09', '2028/01/01', '2021/07/21', '2023/04/10', '2022/07/09', '2023/04/22', '2024/02/08', '2027/11/01', '2025/11/01', '2023/07/19', '2026/08/26', '2022/10/08', '2023/06/12', '2027/06/12', '2020/08/20', '2027/03/28', '2025/10/31', '2026/02/25', '2027/04/09', '2027/02/08', '2024/03/29', '2023/06/29', '2027/05/18', '2023/08/21', '2028/05/06', '2022/04/15', '2022/05/01', '2021/02/12', '2027/12/08', '2022/10/31', '2022/02/01', '2028/02/25', '2020/03/22', '2020/12/08', '2021/04/01', '2021/08/21', '2020/07/31', '2023/12/24', '2028/01/28', '2020/04/11', '2021/01/01', '2028/11/01', '2028/02/27', '2026/05/28', '2026/11/01', '2022/07/10', '2022/08/21', '2021/08/10', '2024/11/30', '2025/04/17', '2022/05/09', '2026/11/30', '2028/04/13', '2023/02/24', '2026/04/05', '2025/01/27', '2027/03/26', '2027/08/21', '2027/12/25', '2020/12/24', '2028/08/21', '2024/12/24', '2024/12/30', '2027/11/02', '2027/01/06', '2024/03/30', '2028/01/26', '2021/05/01', '2023/04/07', '2022/12/26', '2024/04/10', '2025/04/09', '2023/12/25', '2022/12/08', '2020/05/01', '2025/04/20', '2025/12/25', '2021/04/09', '2020/04/12', '2022/07/30', '2022/06/12', '2023/01/01', '2023/02/25', '2026/04/03', '2025/03/02', '2028/04/14', '2027/08/30', '2022/04/16', '2027/11/30', '2026/12/25', '2023/11/30', '2021/11/01', '2024/02/09', '2028/09/03', '2025/03/31', '2026/03/20', '2023/04/06', '2023/10/30', '2025/02/25', '2023/01/02', '2026/12/24', '2028/04/09', '2023/01/22', '2028/08/03', '2020/08/21', '2023/03/23', '2024/07/07', '2022/12/25', '2023/09/27', '2023/11/27', '2024/08/21', '2021/02/25', '2022/05/02', '2023/04/08', '2025/04/19', '2025/01/01', '2022/09/03', '2025/06/12', '2026/02/18', '2021/07/20', '2027/12/31', '2028/04/16', '2027/03/10', '2027/03/25', '2020/06/12', '2027/06/06', '2027/01/01', '2027/12/30', '2020/11/02', '2022/11/01', '2020/05/24', '2021/10/18', '2020/08/31', '2025/08/21', '2024/01/01', '2023/04/09', '2026/12/31', '2028/06/12', '2027/02/06', '2023/11/01', '2023/02/18', '2024/02/10', '2026/11/02', '2027/05/17', '2027/08/15', '2024/06/12', '2020/02/25', '2021/11/30', '2024/09/16', '2026/02/17', '2028/05/01', '2021/03/11', '2026/04/02', '2020/10/29', '2026/06/12', '2024/04/09', '2027/12/24', '2024/03/31', '2025/11/30', '2023/12/31', '2020/12/30', '2021/04/02', '2026/04/04', '2023/04/21', '2028/05/05', '2021/05/13', '2020/04/10', '2023/06/28', '2022/04/14', '2020/05/25', '2023/11/02', '2023/09/03'}






def read_csv(filename):
    if filename.lower().endswith('.csv'):
        with open(filename, newline='', encoding='utf-8') as csvfile:
            return list(csv.reader(csvfile))
    elif filename.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(filename)
        return [df.columns.tolist()] + df.values.tolist()
    else:
        raise ValueError("Unsupported file type. Please upload a .csv or .xlsx/.xls file.")





def adjust_to_next_business_day(date_obj, holidays):
    while date_obj.weekday() >= 5 or date_obj.strftime("%Y/%m/%d") in holidays:
        date_obj += timedelta(days=1)
    return date_obj



def extract(filename):
    add = ""
    for elem in filename:
        if elem == '_':
            break
        else:
            add += elem

    sheet_map = {
        "Loan Data": f"{add}_selected_loans.csv",
    }

    for sheet_name, output_csv in sheet_map.items():
        df = pd.read_excel(filename, sheet_name=sheet_name)
        for col in df.select_dtypes(include=['float', 'int']).columns:
            if (df[col] % 1 == 0).all():
                df[col] = df[col].astype(int)
        df.to_csv(output_csv, index=False)


def edit_merge(file):
    extract(file)
    base_name = file.split('_')[0]
    transactions = f'{base_name}_selected_loans.csv'
    

    
    rows = read_csv(file)
    if not rows or len(rows) < 2:
        return []

    rows = rows[1:]  # Skip header
    head = ''
    head2 = ''
    for row in rows:
        if not pd.isna(row[0]) and row[0] != '':
            head = row[0]
        if pd.isna(row[0]) or row[0] == '':
            row[0] = head

        if not pd.isna(row[1]) and row[1] != '':
            head2 = row[1]
        if pd.isna(row[1]) or row[1] == '':
            row[1] = head2
    return rows


####################################################################
##################################################################





def add_weeks(start, tenor):
    date_obj = datetime.strptime(start, "%Y/%m/%d")
    new_date_obj = date_obj + relativedelta(weeks=tenor)
    return new_date_obj.strftime("%Y/%m/%d")

def add_months(start, tenor):
    date_obj = datetime.strptime(start, "%Y/%m/%d")
    new_date_obj = date_obj + relativedelta(months=tenor)
    return new_date_obj.strftime("%Y/%m/%d")

def generate_due_dates(start_date_str, tenor, freq_days, last):
    start_date = datetime.strptime(start_date_str, "%Y/%m/%d")
    end_date_str = add_months(start_date_str, tenor)
    max_end_date = last
    end_date = min(datetime.strptime(end_date_str, "%Y/%m/%d"),
                   datetime.strptime(max_end_date, "%Y/%m/%d"))
    dates = []
    current = start_date
    while current <= end_date:
        adjusted = adjust_to_next_business_day(current, holidays)
        if adjusted<=end_date:
            dates.append(adjusted.strftime("%Y/%m/%d"))
        current =adjusted + timedelta(days=freq_days)
    return dates

def mapping(payments_with_indices, dates, used,before,after):
    missed = []
    for date in dates:
        date_obj = datetime.strptime(date, "%Y/%m/%d")
        mini = (date_obj - timedelta(days=before)).strftime("%Y/%m/%d") #Window here        
        counter=0
        maxi_obj=date_obj
        while(counter<after):
            maxi_obj += timedelta(days=1)
            if maxi_obj.strftime("%Y/%m/%d") in holidays or maxi_obj.weekday()>=5:
                continue
            counter+=1
        maxi = maxi_obj.strftime("%Y/%m/%d")
              
        matched = False
        matched_date = ''
        for global_idx, (pdate, _),cheque in payments_with_indices:
            if global_idx in used:
                continue
            if mini <= pdate <= maxi:
                used.add(global_idx)
                matched = True
                matched_date = pdate
                break
        if matched:
            missed.append([date, matched_date,global_idx,cheque,"No Default"])
        else:
            missed.append([date, '', '',"DEFAULT",maxi])
    return missed

def default(all_rows, start, tenor, repayment, freq, last, used,before,after):
    if freq == 'Biweekly':
        repayment = round(repayment / 2, 3)
        start = add_weeks(start, 2)
    else:
        start = add_months(start, 0)

    end = min(add_months(start, tenor + 1), last)
    amt = round(repayment)

    filtered = []
    for i, row in enumerate(all_rows):
        try:
            date = row[1]
            amount_str = row[4]
            cheque=row[7]
            if date >= start and date <= end and amount_str != '':
                amount = float(amount_str)
                if round(amount) in [amt, amt + 1, amt - 1]:
                    filtered.append((i, [date, round(amount)],cheque))
        except:
            continue

    if freq == 'Biweekly':
        payment_schedule = generate_due_dates(start, tenor, 14, last)
    else:
        payment_schedule = generate_due_dates(start, tenor, 30, last)

    return mapping(filtered, payment_schedule, used,before,after)

def final(loan_file, transaction_file,before,after):
    used = set()
    res = []
    loan_rows = edit_merge(loan_file)
    loan_rows = list(filter(lambda x: x[-1]==True, loan_rows))
    all_rows = read_csv(transaction_file)[1:]
    last_day = all_rows[-1][1]

    for elem in loan_rows:
        start = elem[0]
        tenor = int(elem[2])
        repayment = float(elem[4])
        freq = elem[5]
        principal = elem[1]
        schedule = default(all_rows, start, tenor, repayment, freq, last_day, used,before,after)
        res.append([start, tenor, principal, repayment, freq, schedule])

    return res
















def export_to_excel(loan_file, transaction_file, before, after, output_path):
    results = final(loan_file, transaction_file, before, after)

    wb = Workbook()
    ws = wb.active
    ws.title = "Repayment Schedule"

    headers = ["Loan Info", "Repayment Schedule", "Repayment Date", "Txn ID", "Cheque No.","Flag"]
    ws.append(headers)

    col_widths = [40, 20, 20, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border=thick_border
        cell.fill=fill_yellow

    current_row = 2
    for loan in results:
        start, tenor, principal, repayment, freq, schedule = loan

        loan_info_text = f"Start Date: {start}\nTenor: {tenor}\nPrinciple: {principal}\nRepayment: {repayment}\nFrequency: {freq}"
        repayment_count = len(schedule)
        end_row = current_row + repayment_count - 1
        merge_range = f"A{current_row}:A{end_row}"
        ws.merge_cells(merge_range)
        cell = ws[f"A{current_row}"]
        cell.value = loan_info_text
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border=thick_border
        for entry in schedule:
            repayment_schedule, repayment_date, txn_id, cheque, flag = entry

            fill = green_fill if flag == "No Default" else red_fill

            
            ws.cell(row=current_row, column=2, value=repayment_schedule)
            ws.cell(row=current_row, column=3, value=repayment_date)
            ws.cell(row=current_row, column=4, value=txn_id)
            ws.cell(row=current_row, column=5, value=cheque)
            ws.cell(row=current_row, column=6, value=flag)

            
            for col in range(2, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill
            current_row += 1

    wb.save(output_path)
    print(f"Excel file saved as '{output_path}'")






def export_custom_results_to_excel(results, output_path):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Repayment Schedule"

    headers = ["Loan Info", "Repayment Schedule", "Repayment Date", "Cheque No." ,"Txn ID", "Flag"]
    ws.append(headers)

    col_widths = [40, 20, 20, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill=fill_yellow
        cell.border=thick_border
    current_row = 2
    for loan in results:
        start, tenor, principal, repayment, freq, cheque,schedule = loan

        loan_info_text = f"Start Date: {start}\nTenor: {tenor}\nPrinciple: {principal}\nRepayment: {repayment}\nFrequency: {freq}"
        repayment_count = len(schedule)
        end_row = current_row + repayment_count - 1
        merge_range = f"A{current_row}:A{end_row}"
        ws.merge_cells(merge_range)
        cell = ws[f"A{current_row}"]
        cell.value = loan_info_text
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border=thick_border
        for entry in schedule:
            repayment_schedule, repayment_date, txn_id, cheque, flag = entry

            fill = green_fill if flag == "No Default" else red_fill

            
            ws.cell(row=current_row, column=2, value=repayment_schedule)
            ws.cell(row=current_row, column=3, value=repayment_date)
            ws.cell(row=current_row, column=4, value=txn_id)
            ws.cell(row=current_row, column=5, value=cheque)
            ws.cell(row=current_row, column=6, value=flag)

            
            for col in range(2, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill
            current_row += 1

    wb.save(output_path)
    print(f"Excel file saved as '{output_path}'")

def manual(start,tenor,repayment,freq,transaction_file,before,after,output_path):
    used = set()
    all_rows=read_csv(transaction_file)[1:]
    last_day=all_rows[-1][1]
    schedule=default(all_rows,start,tenor,repayment,freq,last_day,used,before,after)

    wb=Workbook()
    ws=wb.active
    ws.title='Repayment Schedule'

    
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFACD")

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers=["Loan Info","Repayment Schedule","Repayment Date","Txn ID","Cheque No.","Flag"]
    ws.append(headers)
    for col, header in enumerate(headers,1):
        cell=ws.cell(row=1,column=col,value=header)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    loan_info = f"Start Date: {start}\nTenor: {tenor}\nRepayment: {repayment}\nFrequency: {freq}"
    start_row = 2
    end_row = start_row + len(schedule) - 1
    ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
    cell = ws.cell(row=start_row, column=1, value=loan_info)
    cell.alignment = wrap
    cell.border = border

    for i, row in enumerate(schedule, start=start_row):
        schedule_date, repay_date, txn_id, cheque, flag = row
        fill = green_fill if flag == "No Default" else red_fill
        row_values = [schedule_date, repay_date, txn_id, cheque, flag]

        for j, val in enumerate(row_values, start=2):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    # Set column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Saved manual preview to {output_path}")
    return output_path, schedule

    
    














